from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .models import DeviceRecord, EquipmentGroup, ProjectDefaults


VAV_SD_PARAMETERS = [
    ("SA-AREA", "AV3111", "Default Value", "sa_area"),
    ("SA-KFACTOR", "AV3112", "Default Value", "sa_kfactor"),
    ("CLG-MAXFLOW", "AV3108", "Default Value", "clg_maxflow"),
    ("CLGOCC-MINFLOW", "AV3109", "Default Value", "clg_minflow"),
    ("HTGOCC-MINFLOW", "AV3110", "Default Value", "htg_minflow"),
]


def generate_equipment_name(prefix: str, separator: str, number: int, digits: int) -> str:
    prefix = prefix.strip()
    separator = separator
    if not prefix:
        raise ValueError("Equipment prefix cannot be blank.")
    if digits < 1:
        raise ValueError("Digits must be at least 1.")
    return f"{prefix}{separator}{number:0{digits}d}"


def build_device_description(device_name: str, leaf_space: str, room_number: str) -> str:
    device_name = device_name.strip()
    leaf_space = leaf_space.strip()
    room_number = room_number.strip()

    if not leaf_space:
        return device_name

    description = f"{device_name} - {leaf_space}"
    if room_number:
        description += f" Rm {room_number}"
    return description


def _engine_suffix(engine_name: str) -> str:
    match = re.search(r"(\d{2})\s*$", engine_name.strip())
    if not match:
        raise ValueError("Engine name must end in two digits (example: S3-SNE03).")
    return match.group(1)


def _trunk_suffix(trunk_name: str) -> str:
    match = re.search(r"(\d)\s*$", trunk_name.strip())
    if not match:
        raise ValueError("Trunk name must end in a digit (example: FC-1).")
    return match.group(1)


def generate_bacnet_instance(
    engine_name: str,
    trunk_name: str,
    mac_address: int | None,
    ip_controller_number: int | None,
) -> int:
    """Reproduce the project convention used by the supplied RAC scratchpad.

    Excel convention:
      CONCAT("1000" + RIGHT(engine,2), RIGHT(trunk,1), "00") + MAC + IP_Controller_Number

    MAC and IP controller number are mutually exclusive.
    """
    if mac_address is not None and ip_controller_number is not None:
        raise ValueError("MAC and IP Controller Number cannot both have values.")

    address = mac_address if mac_address is not None else ip_controller_number
    if address is None:
        raise ValueError("A MAC address or IP Controller Number is required to generate an instance.")

    engine = int(_engine_suffix(engine_name))
    trunk = _trunk_suffix(trunk_name)

    base_prefix = 1000 + engine
    base = int(f"{base_prefix}{trunk}00")
    return base + int(address)


def generate_fqr(engine_name: str, trunk_name: str, controller_part: str, instance: int) -> str:
    """Reproduce the FQR convention from the supplied workbook."""
    engine = _engine_suffix(engine_name)
    trunk = _trunk_suffix(trunk_name)
    controller_part = controller_part.strip()
    if len(controller_part) < 5:
        raise ValueError("Controller Part # is too short to generate an FQR.")

    controller_code = controller_part[3:5]
    return f"{engine}{trunk}{controller_code}0{str(instance)[-2:].zfill(2)}"


def infer_heat_flags(controller_template: str) -> tuple[bool | None, bool | None]:
    template = controller_template.strip()
    if not template:
        return None, None
    box_heat = template != "VAV-CLG"
    supplemental = template == "VAV-RAD"
    return box_heat, supplemental


def load_manufacturer_data(template_path: str | Path) -> dict[str, dict[int, tuple[float | None, float | None]]]:
    """Read the manufacturer Area/K Factor database from the supplied template.

    The workbook stores manufacturers in pairs of columns beginning at D/E,
    with names on row 25, labels on row 27, and duct sizes in B28:B45.
    """
    wb = load_workbook(template_path, data_only=True, read_only=True)
    ws = wb["Manufacturer"]

    manufacturers: dict[str, dict[int, tuple[float | None, float | None]]] = {}
    for area_col in range(4, 38, 2):  # D through AK, two columns per manufacturer
        name = ws.cell(row=25, column=area_col).value
        if not name:
            continue
        display_name = str(name).split("\n", 1)[0].strip()
        records: dict[int, tuple[float | None, float | None]] = {}
        for row in range(28, 46):
            size = ws.cell(row=row, column=2).value
            if size is None:
                continue
            try:
                size_int = int(size)
            except (TypeError, ValueError):
                continue

            area = ws.cell(row=row, column=area_col).value
            kfactor = ws.cell(row=row, column=area_col + 1).value

            def numeric_or_none(value):
                return float(value) if isinstance(value, (int, float)) else None

            records[size_int] = (numeric_or_none(area), numeric_or_none(kfactor))
        manufacturers[display_name] = records

    wb.close()
    return manufacturers


def manufacturer_lookup(
    manufacturer_data: dict[str, dict[int, tuple[float | None, float | None]]],
    manufacturer: str,
    inlet_size: int | None,
) -> tuple[float | None, float | None]:
    if inlet_size is None:
        return None, None
    return manufacturer_data.get(manufacturer, {}).get(int(inlet_size), (None, None))


def recalculate_record(
    record: DeviceRecord,
    manufacturer_data: dict[str, dict[int, tuple[float | None, float | None]]],
    generate_instance: bool = True,
) -> DeviceRecord:
    record.device_description = build_device_description(
        record.device_name, record.leaf_space, record.room_number
    )

    record.box_heat, record.supplemental_heat = infer_heat_flags(record.controller_template)
    record.sa_area, record.sa_kfactor = manufacturer_lookup(
        manufacturer_data, record.manufacturer, record.inlet_size
    )

    if generate_instance:
        try:
            record.instance = generate_bacnet_instance(
                record.engine_name,
                record.trunk_name,
                record.mac_address,
                record.ip_controller_number,
            )
        except ValueError:
            record.instance = None
    else:
        record.instance = None

    if record.instance is not None:
        try:
            record.fqr = generate_fqr(
                record.engine_name,
                record.trunk_name,
                record.controller_part,
                record.instance,
            )
        except ValueError:
            record.fqr = ""
    else:
        record.fqr = ""

    # Standard single-duct VAV parameters supported by the current first release.
    record.parameters = {
        "SA-AREA": record.sa_area,
        "SA-KFACTOR": record.sa_kfactor,
        "CLG-MAXFLOW": record.clg_maxflow,
        "CLGOCC-MINFLOW": record.clg_minflow,
        "HTGOCC-MINFLOW": record.htg_minflow,
    }
    return record


def generate_group_records(
    defaults: ProjectDefaults,
    group: EquipmentGroup,
    manufacturer_data: dict[str, dict[int, tuple[float | None, float | None]]],
) -> list[DeviceRecord]:
    if group.end < group.start:
        raise ValueError("Ending number must be greater than or equal to the starting number.")
    if group.digits < 1:
        raise ValueError("Digits must be at least 1.")
    if group.start_address < 0:
        raise ValueError("Starting MAC/IP Controller Number cannot be negative.")

    records: list[DeviceRecord] = []
    for offset, number in enumerate(range(group.start, group.end + 1)):
        equipment_name = generate_equipment_name(
            group.equipment_prefix, group.separator, number, group.digits
        )
        device_name = f"{defaults.device_prefix}{equipment_name}"
        address = group.start_address + offset

        record = DeviceRecord(
            site_hierarchy=defaults.site_hierarchy,
            device_name=device_name,
            equipment_name=equipment_name,
            served_by=group.served_by,
            controller_part=defaults.controller_part,
            engine_name=defaults.engine_name,
            trunk_name=defaults.trunk_name,
            equipment_definition=defaults.equipment_definition,
            controller_template=defaults.controller_template,
            dhcp_enabled=defaults.dhcp_enabled if defaults.network_type == "IP" else None,
            subnet_mask=defaults.subnet_mask if defaults.network_type == "IP" else "",
            ip_router=defaults.ip_router if defaults.network_type == "IP" else "",
            manufacturer=group.manufacturer,
            inlet_size=group.inlet_size,
            clg_maxflow=group.clg_maxflow,
            clg_minflow=group.clg_minflow,
            htg_minflow=group.htg_minflow,
        )

        if defaults.network_type == "MSTP":
            record.mac_address = address
        else:
            record.ip_controller_number = address

        recalculate_record(record, manufacturer_data, defaults.generate_instance)
        records.append(record)

    return records


def find_duplicate_values(records: Iterable[DeviceRecord]) -> dict[str, set[object]]:
    """Return duplicate identifiers that are important for RAC uniqueness."""
    fields = {
        "Equipment Name": [r.equipment_name for r in records],
        "Device Name": [r.device_name for r in records],
        "FQR": [r.fqr for r in records if r.fqr],
        "Instance": [r.instance for r in records if r.instance is not None],
    }
    duplicates: dict[str, set[object]] = {}
    for label, values in fields.items():
        seen = set()
        dupes = set()
        for value in values:
            if value in seen:
                dupes.add(value)
            seen.add(value)
        if dupes:
            duplicates[label] = dupes
    return duplicates

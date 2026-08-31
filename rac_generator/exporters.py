from __future__ import annotations

import csv
import shutil
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .logic import VAV_SD_PARAMETERS, build_dependency_levels, ordered_records_for_sct
from .models import DeviceRecord


BASE_RAC_HEADERS = [
    "RAC-4096", "RAC-4112", "RAC-4128", "RAC-4144", "RAC-4160", "RAC-4166",
    "RAC-4170", "RAC-4176", "RAC-4192", "RAC-4208", "RAC-4224", "RAC-4240",
    "RAC-4256", "RAC-4272", "RAC-4288", "RAC-4304", "RAC-4311", "RAC-4320",
    "RAC-4336", "RAC-4352", "RAC-4368", "RAC-4384", "RAC-4400", "RAC-4416",
    "RAC-4432", "RAC-4448",
]

BASE_FIELD_NAMES = [
    "Site/Building/Floor\n(Required)", "Room Number\n(Optional)",
    "Leaf Space (e.g. Room)\n(Required)", "Device Name\n(Required)",
    "Device FQR Reference \n(Required)", "Device Description\n(Optional)",
    "Equipment Name\n(Required)", "Served By Equipment Name\n(Optional)",
    "Controller Part #\n(Optional)", "Engine Name\n(Required)", "Trunk Name\n(Required)",
    "Controller Host Name \n(Future)", "JCI MAC  Address", "IP Controller Number",
    "ZIGBEE PAN Offset", "Instance # (BACoid)", "N2Address", "DHCP Enabled ",
    "IP Address ", "Subnet Mask", "IP Router", "ETH-1\n(Optional)", "ETH-2\n(Optional)",
    "Equipment Definition Name", "Controller Template Name", None,
]

SCT_SETUP_GUIDE = [
    ("1", "Create/open the SCT archive", "The archive must contain the site structure that will receive the field controllers."),
    ("2", "Create Site, Site Director, supervisory devices, and integrations", "The Engine Name and Trunk Name used by the RAC schedule must already exist exactly as named in SCT."),
    ("3", "Create or import Equipment Definitions", "In SCT use Facility > Prepare Rapid Archive > Insert Equipment Definition, or import definitions into the Definitions folder. Johnson Controls recommends creating definitions bottom-up: terminal units, then AHUs, then central plant."),
    ("4", "Create or import Controller Templates", "Create under Configuration > SCT Controller Templates. Link each template to the correct Equipment Definition using the Definition Link."),
    ("5", "Verify Controller Template Definition Links", "Equipment Definition Name in the RAC spreadsheet is reference-only; the functional Equipment Definition relationship comes from the Controller Template Definition Link."),
    ("6", "Confirm any pre-existing serving equipment", "If a Generated device is Served By equipment not generated in this project, list that parent in the app's 'Equipment already in SCT' field and verify it exists in the archive with the exact name."),
    ("7", "Import Rapid Archive files top-down", "Import Level 01 first and Save in the Rapid Archive wizard before Level 02, then continue level-by-level. This allows Served By relationships to resolve to equipment already created."),
    ("8", "Review before Save", "Verify Supervisory Device, Integration, Controller Template, spaces, addresses, instances, CAF parameters, and Served By relationships in the Rapid Archive wizard."),
]


def _record_base_values(record: DeviceRecord) -> list[object]:
    return [
        record.site_hierarchy, record.room_number, record.leaf_space, record.device_name,
        record.fqr, record.device_description, record.equipment_name, record.served_by,
        record.controller_part, record.engine_name, record.trunk_name,
        record.controller_host_name, record.mac_address, record.ip_controller_number,
        record.zigbee_pan_offset, record.instance, record.n2_address, record.dhcp_enabled,
        record.ip_address, record.subnet_mask, record.ip_router, record.eth1, record.eth2,
        record.equipment_definition, record.controller_template,
    ]


def _parameters_in_use(records: Iterable[DeviceRecord]):
    records = list(records)
    used = []
    for name, attr_id, attr_type, field_name in VAV_SD_PARAMETERS:
        if any(getattr(r, field_name) is not None for r in records):
            used.append((name, attr_id, attr_type, field_name))
    return used


def _rac_rows(records: list[DeviceRecord]) -> list[list[object]]:
    params = _parameters_in_use(records)
    total_cols = 25 + max(1, len(params))
    rows: list[list[object]] = []

    row1 = BASE_RAC_HEADERS[:25] + ["RAC-4448"]
    if len(params) > 1:
        row1 += [None] * (len(params) - 1)
    rows.append(row1)
    rows.append(["v3"] + [None] * (total_cols - 1))

    row3 = [None] * total_cols
    row3[0] = "Space Information"
    row3[3] = "Network / Equipment Tree Information"
    row3[9] = "Network Information (MSTP and IP)"
    row3[23] = "Definitions and Templates"
    row3[25] = "Parameters"
    rows.append(row3)

    row4 = BASE_FIELD_NAMES[:25]
    row4 += [p[0] for p in params] if params else [None]
    rows.append(row4)

    row5 = [None] * total_cols
    row6 = [None] * total_cols
    for idx, param in enumerate(params, start=25):
        row5[idx] = param[1]
        row6[idx] = param[2]
    row5[3] = "Attribute ID"
    row6[3] = "Attribute Type"
    rows.extend([row5, row6])

    for record in records:
        values = _record_base_values(record)
        values += [getattr(record, p[3]) for p in params] if params else [None]
        rows.append(values)
    return rows


def export_rac_csv(path: str | Path, records: list[DeviceRecord]) -> None:
    path = Path(path)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        csv.writer(handle).writerows(_rac_rows(records))


def export_staged_rac_csvs(
    directory: str | Path,
    records: list[DeviceRecord],
    existing_equipment: str | Iterable[str] = (),
) -> list[Path]:
    """Export one SCT CSV per dependency level, in the required top-down import order."""
    directory = Path(directory)
    directory.mkdir(parents=True, exist_ok=True)
    levels = build_dependency_levels(records, existing_equipment)
    paths: list[Path] = []
    for number, level in enumerate(levels, start=1):
        path = directory / f"SCT_{number:02d}_Level_{number - 1}.csv"
        export_rac_csv(path, level)
        paths.append(path)
    return paths


def _style_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_setup_guide(wb) -> None:
    if "SCT Setup Guide" in wb.sheetnames:
        del wb["SCT Setup Guide"]
    ws = wb.create_sheet("SCT Setup Guide", 0)
    headers = ["Step", "What must exist / what to do", "Why / SCT guidance"]
    for col, header in enumerate(headers, start=1):
        _style_header(ws.cell(1, col, header))
    for row, values in enumerate(SCT_SETUP_GUIDE, start=2):
        for col, value in enumerate(values, start=1):
            ws.cell(row, col, value).alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 95
    ws.freeze_panes = "A2"


def _add_import_plan(wb, levels: list[list[DeviceRecord]], existing_equipment: str | Iterable[str]) -> None:
    if "SCT Import Plan" in wb.sheetnames:
        del wb["SCT Import Plan"]
    ws = wb.create_sheet("SCT Import Plan", 1)
    headers = ["Import Order", "Equipment Name", "Device Name", "Served By", "Engine", "Trunk", "Instruction"]
    for col, header in enumerate(headers, start=1):
        _style_header(ws.cell(1, col, header))
    row = 2
    for level_index, level in enumerate(levels, start=1):
        for record in level:
            values = [
                level_index, record.equipment_name, record.device_name, record.served_by,
                record.engine_name, record.trunk_name,
                f"Import SCT_{level_index:02d}_Level_{level_index - 1}.csv, then Save in Rapid Archive before the next level.",
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row, col, value)
            row += 1
    existing = existing_equipment if isinstance(existing_equipment, str) else ", ".join(existing_equipment)
    ws.cell(row + 1, 1, "Pre-existing equipment declared in app:")
    ws.cell(row + 1, 2, existing or "None")
    for col, width in enumerate([14, 24, 26, 30, 20, 16, 70], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def export_rac_workbook(
    output_path: str | Path,
    template_path: str | Path,
    records: list[DeviceRecord],
    existing_equipment: str | Iterable[str] = (),
) -> None:
    output_path = Path(output_path)
    template_path = Path(template_path)
    shutil.copy2(template_path, output_path)

    ordered = ordered_records_for_sct(records, existing_equipment)
    levels = build_dependency_levels(records, existing_equipment)
    wb = load_workbook(output_path)
    _add_setup_guide(wb)
    _add_import_plan(wb, levels, existing_equipment)

    rac = wb["Rapid Archive Schedule"]
    params = _parameters_in_use(ordered)
    total_cols = 25 + max(1, len(params))

    for row in rac.iter_rows(min_row=7, max_row=max(rac.max_row, 7), min_col=1, max_col=max(rac.max_column, total_cols)):
        for cell in row:
            cell.value = None

    rac.cell(1, 26).value = "RAC-4448"
    rac.cell(3, 26).value = "Parameters"
    for idx, param in enumerate(params, start=26):
        rac.cell(1, idx).value = "RAC-4448" if idx == 26 else None
        rac.cell(4, idx).value = param[0]
        rac.cell(5, idx).value = param[1]
        rac.cell(6, idx).value = param[2]

    for row_index, record in enumerate(ordered, start=7):
        for col, value in enumerate(_record_base_values(record), start=1):
            rac.cell(row_index, col).value = value
        for param_index, param in enumerate(params, start=26):
            rac.cell(row_index, param_index).value = getattr(record, param[3])

    if "Generated Scratchpad" in wb.sheetnames:
        del wb["Generated Scratchpad"]
    scratch = wb.create_sheet("Generated Scratchpad")
    scratch_headers = [
        "Site/Building/Floor", "Room Number", "Leaf Space", "Device Name", "Device FQR Reference",
        "Device Description", "Equipment Name", "Mechanical Drawing", "Served By Equipment Name",
        "JCI Ctrl Dwg No.", "Controller Part #", "Engine Name", "Trunk Name", "Controller Host Name",
        "JCI MAC Address", "IP Controller Number", "ZigBee PAN Offset", "Instance # (BACoid)",
        "N2 Address", "DHCP Enabled", "IP Address", "Subnet Mask", "IP Router", "ETH-1", "ETH-2",
        "Equipment Definition Name (Reference Only)", "Controller Template Name", "Sensor Code No.",
        "Box Heat", "Supplemental Heat", "Inlet Size", "SA-AREA", "SA-KFACTOR", "CLG-MAXFLOW",
        "CLGOCC-MINFLOW", "HTGOCC-MINFLOW", "Comments",
    ]
    for col, header in enumerate(scratch_headers, start=1):
        _style_header(scratch.cell(1, col, header))

    for row_index, record in enumerate(ordered, start=2):
        values = [
            record.site_hierarchy, record.room_number, record.leaf_space, record.device_name, record.fqr,
            record.device_description, record.equipment_name, record.mechanical_drawing, record.served_by,
            record.jci_ctrl_dwg_no, record.controller_part, record.engine_name, record.trunk_name,
            record.controller_host_name, record.mac_address, record.ip_controller_number,
            record.zigbee_pan_offset, record.instance, record.n2_address, record.dhcp_enabled,
            record.ip_address, record.subnet_mask, record.ip_router, record.eth1, record.eth2,
            record.equipment_definition, record.controller_template, record.sensor_code_no, record.box_heat,
            record.supplemental_heat, record.inlet_size, record.sa_area, record.sa_kfactor,
            record.clg_maxflow, record.clg_minflow, record.htg_minflow, record.comments,
        ]
        for col, value in enumerate(values, start=1):
            scratch.cell(row_index, col).value = value

    scratch.freeze_panes = "A2"
    scratch.auto_filter.ref = f"A1:{get_column_letter(len(scratch_headers))}{max(1, len(ordered) + 1)}"
    for col in range(1, len(scratch_headers) + 1):
        scratch.column_dimensions[get_column_letter(col)].width = 24 if col in (1, 3, 4, 5, 6, 7, 9, 26, 27, 37) else 18
    scratch.row_dimensions[1].height = 45
    wb.save(output_path)
